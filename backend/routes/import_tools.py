"""工装模板下载与Excel批量导入"""
from flask import Blueprint, request, jsonify, send_file, session
from models import db, Tool
from routes.auth import login_required, edit_required, add_log
from datetime import datetime
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import_bp = Blueprint('import', __name__)

# 模板列定义：(表头名, 对应字段, 是否必填, 说明)
TEMPLATE_COLUMNS = [
    ('编号', 'code', True, '必填，且系统内唯一'),
    ('图号', 'drawing_no', False, '可选'),
    ('名称', 'name', True, '必填'),
    ('规格型号', 'spec', False, '可选'),
    ('类别', 'category', False, '可选，如：量具/刀具/夹具'),
    ('等级', 'level', False, '可选，A/B/C，默认A'),
    ('使用分厂', 'factory', False, '可选'),
    ('使用班组', 'team', False, '可选'),
    ('领用人', 'receiver', False, '可选'),
    ('状态', 'status', False, '可选，默认"在库"'),
    ('完工日期', 'purchase_date', False, '可选，格式YYYY-MM-DD'),
    ('下次检定日期', 'next_inspection_date', False, '可选，格式YYYY-MM-DD'),
    ('备注', 'remark', False, '可选'),
]

DATE_FIELDS = ('purchase_date', 'next_inspection_date')

# 示例行哨兵：模板示例行的备注含此字符串，导入时自动跳过，避免用户忘记删示例行导致垃圾数据
EXAMPLE_SENTINEL = 'TEMPLATE_EXAMPLE_DO_NOT_IMPORT'


def _header_style(ws, ncols):
    font = Font(bold=True, color='FFFFFF', size=11)
    fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = thin


@import_bp.route('/template', methods=['GET'])
@login_required
def download_template():
    """下载工装导入模板（含示例行）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '工装导入模板'

    # 表头
    ws.append([c[0] for c in TEMPLATE_COLUMNS])
    _header_style(ws, len(TEMPLATE_COLUMNS))

    # 示例数据（第2行，浅灰提示）
    example = ['GZ2026072001', 'DWG-001', '卡箍接头φ51转φ38焊接', 'φ51→φ38',
               '焊接类', 'A', '一厂', '一班', '张三', '在库',
               '2026-07-20', '2026-10-20', EXAMPLE_SENTINEL]
    ws.append(example)
    for col in range(1, len(TEMPLATE_COLUMNS) + 1):
        ws.cell(row=2, column=col).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws.cell(row=2, column=col).alignment = Alignment(horizontal='center')

    # 列宽
    widths = [14, 14, 22, 16, 12, 6, 12, 12, 10, 8, 14, 16, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='工装导入模板.xlsx'
    )


@import_bp.route('/tools', methods=['POST'])
@edit_required
def import_tools():
    """上传Excel批量导入工装"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '未找到上传文件'})

    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'msg': '仅支持 .xlsx / .xls 文件'})

    try:
        wb = openpyxl.load_workbook(f.stream, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'code': 400, 'msg': f'Excel解析失败：{str(e)}'})

    ws = wb.active

    # 读取表头（第1行）
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({'code': 400, 'msg': '文件为空'})
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    # 建立 表头名 → 列索引 映射
    col_index = {}
    for i, h in enumerate(headers):
        for col_def in TEMPLATE_COLUMNS:
            if h == col_def[0]:
                col_index[col_def[1]] = i
                break

    # 检查必填列是否存在
    for col_def in TEMPLATE_COLUMNS:
        if col_def[2] and col_def[1] not in col_index:
            return jsonify({'code': 400, 'msg': f'缺少必填列：{col_def[0]}'})

    success = 0
    skipped = 0
    errors = []

    for r_idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        # 取字段值
        def get_val(field):
            idx = col_index.get(field)
            if idx is None or idx >= len(row):
                return ''
            v = row[idx]
            if v is None:
                return ''
            return str(v).strip()

        code = get_val('code')
        name = get_val('name')

        # 空行跳过（编号和名称都为空）
        if not code and not name:
            continue

        # 跳过模板示例行（哨兵标记）
        if get_val('remark') == EXAMPLE_SENTINEL:
            continue

        if not code:
            errors.append(f'第{r_idx}行：编号不能为空')
            continue
        if not name:
            errors.append(f'第{r_idx}行（编号{code}）：名称不能为空')
            continue

        # 编号重复检查
        if Tool.query.filter_by(code=code).first():
            skipped += 1
            continue

        tool = Tool(
            code=code,
            drawing_no=get_val('drawing_no'),
            name=name,
            spec=get_val('spec'),
            category=get_val('category'),
            level=get_val('level') or 'A',
            factory=get_val('factory'),
            team=get_val('team'),
            receiver=get_val('receiver'),
            status=get_val('status') or '在库',
            remark=get_val('remark'),
        )

        # 日期解析
        ok = True
        for df in DATE_FIELDS:
            raw = get_val(df)
            if raw:
                try:
                    d = datetime.strptime(raw, '%Y-%m-%d').date()
                    setattr(tool, df, d)
                except ValueError:
                    errors.append(f'第{r_idx}行（编号{code}）：{df}格式应为YYYY-MM-DD，当前值「{raw}」')
                    ok = False
                    break
        if not ok:
            continue

        db.session.add(tool)
        success += 1

    # 提交
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'msg': f'数据库写入失败：{str(e)}'})

    if success > 0:
        add_log('批量导入工装', f'成功导入 {success} 条，跳过 {skipped} 条，错误 {len(errors)} 条')

    return jsonify({
        'code': 200,
        'msg': f'导入完成：成功 {success} 条，跳过（编号重复）{skipped} 条，错误 {len(errors)} 条',
        'data': {
            'success': success,
            'skipped': skipped,
            'error_count': len(errors),
            'errors': errors[:50],  # 最多返回50条错误明细
        }
    })
