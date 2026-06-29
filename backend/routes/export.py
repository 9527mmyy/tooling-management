from flask import Blueprint, request, jsonify, session, send_file
from models import db, Tool, BorrowRecord, Inspection
from routes.auth import login_required
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

export_bp = Blueprint('export', __name__)


def style_header(ws, headers):
    """设置表头样式"""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_row(ws, row_num, col_count):
    """设置数据行样式"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    even_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if row_num % 2 == 0:
            cell.fill = even_fill


def fmt_date(dt):
    """格式化日期，兼容datetime和date"""
    if dt is None:
        return ''
    if isinstance(dt, datetime):
        return dt.strftime('%Y年%m月%d日 %H:%M')
    return dt.strftime('%Y年%m月%d日')


# ─────────────────────────────────────────────
# 工装台账导出
# ─────────────────────────────────────────────
@export_bp.route('/tools', methods=['GET'])
@login_required
def export_tools():
    """导出工装台账Excel"""
    keyword = request.args.get('keyword', '').strip()
    status = request.args.get('status', '').strip()
    factory = request.args.get('factory', '').strip()
    team = request.args.get('team', '').strip()

    query = Tool.query
    if keyword:
        query = query.filter(
            db.or_(
                Tool.code.contains(keyword),
                Tool.name.contains(keyword),
                Tool.spec.contains(keyword)
            )
        )
    if status:
        query = query.filter_by(status=status)
    if factory:
        query = query.filter(Tool.factory.contains(factory))
    if team:
        query = query.filter(Tool.team.contains(team))

    tools = query.order_by(Tool.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '工装台账'

    headers = ['序号', '编号', '图号', '名称', '规格型号', '类别', '等级', '使用分厂', '使用班组', '领用人', '状态', '完工日期', '下次检定日期', '备注']
    ws.append(headers)
    style_header(ws, headers)

    for i, tool in enumerate(tools, 1):
        ws.append([
            i,
            tool.code,
            tool.drawing_no or '',
            tool.name,
            tool.spec or '',
            tool.category or '',
            tool.level or 'A',
            tool.factory or '',
            tool.team or '',
            tool.receiver or '',
            tool.status,
            tool.purchase_date.strftime('%Y年%m月%d日') if tool.purchase_date else '',
            tool.next_inspection_date.strftime('%Y年%m月%d日') if tool.next_inspection_date else '',
            tool.remark or '',
        ])
        style_data_row(ws, i + 1, len(headers))

    # 列宽自适应
    col_widths = [6, 14, 14, 18, 18, 10, 6, 14, 14, 10, 8, 16, 16, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30

    filename = f'工装台账_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)


# ─────────────────────────────────────────────
# 借用记录导出
# ─────────────────────────────────────────────
@export_bp.route('/borrows', methods=['GET'])
@login_required
def export_borrows():
    """导出借用记录Excel"""
    status = request.args.get('status', '').strip()
    borrower = request.args.get('borrower', '').strip()
    factory = request.args.get('factory', '').strip()
    team = request.args.get('team', '').strip()

    query = BorrowRecord.query
    if status:
        query = query.filter_by(status=status)
    if borrower:
        query = query.filter(BorrowRecord.borrower.contains(borrower))
    if factory:
        query = query.filter(BorrowRecord.factory.contains(factory))
    if team:
        query = query.filter(BorrowRecord.team.contains(team))

    records = query.order_by(BorrowRecord.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '借用记录'

    headers = ['序号', '工装编号', '工装名称', '使用分厂', '使用班组', '领用人', '借用人', '借用时间', '归还时间', '状态']
    ws.append(headers)
    style_header(ws, headers)

    for i, rec in enumerate(records, 1):
        ws.append([
            i,
            rec.tool.code if rec.tool else '',
            rec.tool.name if rec.tool else '',
            rec.factory or '',
            rec.team or '',
            rec.receiver or '',
            rec.borrower,
            fmt_date(rec.borrow_time),
            fmt_date(rec.return_time),
            rec.status,
        ])
        style_data_row(ws, i + 1, len(headers))

    col_widths = [6, 14, 18, 14, 14, 10, 10, 20, 20, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30

    filename = f'借用记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)


# ─────────────────────────────────────────────
# 检定记录导出
# ─────────────────────────────────────────────
@export_bp.route('/inspections', methods=['GET'])
@login_required
def export_inspections():
    """导出检定记录Excel"""
    tool_id = request.args.get('tool_id', type=int)
    result = request.args.get('result', '').strip()

    query = Inspection.query
    if tool_id:
        query = query.filter_by(tool_id=tool_id)
    if result:
        query = query.filter_by(result=result)

    records = query.order_by(Inspection.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '检定记录'

    headers = ['序号', '工装编号', '工装名称', '检定日期', '检定结果', '下次检定日期', '检定员', '备注']
    ws.append(headers)
    style_header(ws, headers)

    for i, rec in enumerate(records, 1):
        ws.append([
            i,
            rec.tool.code if rec.tool else '',
            rec.tool.name if rec.tool else '',
            rec.inspect_date.strftime('%Y年%m月%d日') if rec.inspect_date else '',
            rec.result,
            rec.next_date.strftime('%Y年%m月%d日') if rec.next_date else '',
            rec.inspector or '',
            rec.remark or '',
        ])
        style_data_row(ws, i + 1, len(headers))

    col_widths = [6, 14, 18, 16, 10, 16, 12, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30

    filename = f'检定记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
